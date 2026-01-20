import logging 
import os
import time

from io import BytesIO

from openpyxl import load_workbook
from tabulate import tabulate

from .base_chunker import BaseChunker
from dependencies import get_config

app_config_client = get_config()

class SpreadsheetChunker(BaseChunker):
    """
    SpreadsheetChunker processes and chunks spreadsheet content, such as Excel files, into manageable pieces for analysis and summarization. 
    It handles both chunking by rows or sheets, allowing users to specify whether to include header rows in each chunk, and ensures that 
    the content size does not exceed a specified token limit.

    The class supports the following operations:
    - Converts spreadsheets into chunkable content.
    - Provides options to chunk either by row or by sheet.
    - Includes optional header rows in chunks.
    - Summarizes large sheets if the content exceeds the maximum chunk size.
    
    Attributes:
    -----------
    max_chunk_size (int): Maximum allowed size of each chunk in tokens.
    chunking_by_row (bool): Whether to chunk by row instead of by sheet.
    include_header_in_chunks (bool): Whether to include header rows in each row-based chunk.
    document_content (str): Processed spreadsheet content ready for chunking.

    Methods:
    --------
    - get_chunks(): Splits the spreadsheet content into manageable chunks, based on the configuration.
    - _spreadsheet_process(): Extracts and processes data from each sheet, including summaries if necessary.
    - _get_sheet_data(sheet): Retrieves data and headers from the given sheet, handling empty cells.
    - _clean_markdown_table(table_str): Cleans up Markdown table strings by removing excessive whitespace.
    """

    def __init__(self, data, max_chunk_size=None, chunking_by_row=None, include_header_in_chunks=None):
        """
        Initializes the SpreadsheetChunker with the provided data and environment configurations.
        
        Args:
            data (str): The spreadsheet content to be chunked.
            max_chunk_size (int, optional): Maximum allowed size of each chunk in tokens. Defaults to an environment variable 'SPREADSHEET_CHUNKING_NUM_TOKENS' or 0 if not set.
            chunking_by_row (bool, optional): Whether to chunk by row instead of by sheet. Defaults to an environment variable 'CHUNKING_BY_ROW' or False.
            include_header_in_chunks (bool, optional): Whether to include the header row in each chunk if chunking by row. Defaults to 'INCLUDE_HEADER_IN_CHUNKS' environment variable or False.
        """
        super().__init__(data)
        
        if max_chunk_size is None:
            self.max_chunk_size = int(app_config_client.get("SPREADSHEET_CHUNKING_NUM_TOKENS", 0))
        else:
            self.max_chunk_size = int(max_chunk_size)
        
        if chunking_by_row is None:
            chunking_env = app_config_client.get("SPREADSHEET_CHUNKING_BY_ROW", "false").lower()
            self.chunking_by_row = chunking_env in ["true", "1", "yes"]
        else:
            self.chunking_by_row = bool(chunking_by_row)
        
        if include_header_in_chunks is None:
            include_header_env = app_config_client.get("SPREADSHEET_CHUNKING_BY_ROW_INCLUDE_HEADER", "false").lower()
            self.include_header_in_chunks = include_header_env in ["true", "1", "yes"]
        else:
            self.include_header_in_chunks = bool(include_header_in_chunks)

    def get_chunks(self):
        """
        Splits the spreadsheet content into smaller chunks. Depending on the configuration, chunks can be created by sheet or by row.
        - If chunking by sheet, the method summarizes content that exceeds the maximum chunk size.
        - If chunking by row, each row is processed into its own chunk, optionally including the header row.
        
        Returns:
            List[dict]: A list of dictionaries representing the chunks created from the spreadsheet.
        """
        return list(self.iter_chunks())

    def iter_chunks(self):
        """Yield chunks one-by-one to avoid buffering all chunks/vectors in memory."""
        logging.info(f"[spreadsheet_chunker][{self.filename}][iter_chunks] Running iter_chunks.")
        total_start_time = time.time()

        blob_stream = BytesIO(self.document_bytes)
        workbook = load_workbook(blob_stream, data_only=True)
        logging.info(
            f"[spreadsheet_chunker][{self.filename}][iter_chunks] Workbook has {len(workbook.sheetnames)} sheets"
        )

        chunk_id = 0
        if not self.chunking_by_row:
            # Original behavior: Chunk per sheet
            for sheet_name in workbook.sheetnames:
                start_time = time.time()
                current_chunk_id = chunk_id
                sheet = workbook[sheet_name]
                logging.debug(
                    f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                    f"Starting processing chunk {current_chunk_id} (sheet)."
                )

                data, headers = self._get_sheet_data(sheet)
                table_content = tabulate(data, headers=headers, tablefmt="grid")
                table_content = self._clean_markdown_table(table_content)
                table_tokens = self.token_estimator.estimate_tokens(table_content)

                prompt = (
                    f"Summarize the table with data in it, by understanding the information clearly.\n "
                    f"table_data:{table_content}"
                )
                summary = self.aoai_client.get_completion(prompt, max_tokens=2048)

                if self.max_chunk_size > 0 and table_tokens > self.max_chunk_size:
                    logging.info(
                        f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                        f"Table has {table_tokens} tokens. Max tokens is {self.max_chunk_size}. Using summary."
                    )
                    table_content = summary

                chunk_dict = self._create_chunk(
                    chunk_id=current_chunk_id,
                    content=table_content,
                    summary=summary,
                    embedding_text=summary if summary else table_content,
                    title=sheet_name,
                )
                yield chunk_dict
                chunk_id += 1
                elapsed_time = time.time() - start_time
                logging.debug(
                    f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                    f"Processed chunk {current_chunk_id} in {elapsed_time:.2f} seconds."
                )
        else:
            # New behavior: Chunk per row (streaming)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logging.info(
                    f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                    "Starting row-wise chunking."
                )

                try:
                    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                except StopIteration:
                    continue
                headers = ["" if v is None else str(v) for v in (header_row or [])]

                for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    row = ["" if v is None else str(v) for v in (row_values or [])]
                    if not any(cell.strip() for cell in row):
                        continue

                    start_time = time.time()
                    current_chunk_id = chunk_id
                    logging.debug(
                        f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                        f"Processing chunk {current_chunk_id} for row {row_index}."
                    )

                    if self.include_header_in_chunks:
                        table = tabulate([headers, row], headers="firstrow", tablefmt="github")
                    else:
                        table = tabulate([row], headers=headers, tablefmt="github")

                    table = self._clean_markdown_table(table)
                    content = table
                    embedding_text = self._row_to_embedding_text(
                        headers=headers,
                        row=row,
                        sheet_name=sheet_name,
                        row_index=row_index,
                        include_header_in_embedding=self.include_header_in_chunks,
                    )

                    if self.max_chunk_size and self.max_chunk_size > 0:
                        content_tokens = self.token_estimator.estimate_tokens(content)
                        if content_tokens > self.max_chunk_size:
                            logging.info(
                                f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                                f"Row content has {content_tokens} tokens. Max tokens is {self.max_chunk_size}. Truncating."
                            )
                            content = self._truncate_chunk(content)
                        embed_tokens = self.token_estimator.estimate_tokens(embedding_text)
                        if embed_tokens > self.max_chunk_size:
                            embedding_text = self._truncate_chunk(embedding_text)

                    chunk_dict = self._create_chunk(
                        chunk_id=current_chunk_id,
                        content=content,
                        summary="",
                        embedding_text=embedding_text,
                        title=f"{sheet_name} - Row {row_index}",
                    )
                    yield chunk_dict
                    chunk_id += 1
                    elapsed_time = time.time() - start_time
                    logging.debug(
                        f"[spreadsheet_chunker][{self.filename}][iter_chunks][{sheet_name}] "
                        f"Processed chunk {current_chunk_id} in {elapsed_time:.2f} seconds."
                    )

        total_elapsed_time = time.time() - total_start_time
        logging.debug(
            f"[spreadsheet_chunker][{self.filename}][iter_chunks] Finished iter_chunks in {total_elapsed_time:.2f} seconds."
        )

    def _spreadsheet_process(self):
        """
        Extracts and processes each sheet from the spreadsheet, converting the content into Markdown table format. 
        If chunking by sheet, a summary is generated if the sheet's content exceeds the maximum token size.

        Returns:
            List[dict]: A list of dictionaries, where each dictionary contains sheet metadata, headers, rows, table content, and a summary if applicable.
        """
        logging.debug(f"[spreadsheet_chunker][{self.filename}][spreadsheet_process] Starting blob download.")        
        blob_data = self.document_bytes
        blob_stream = BytesIO(blob_data)
        logging.debug(f"[spreadsheet_chunker][{self.filename}][spreadsheet_process] Starting openpyxl load_workbook.")                    
        workbook = load_workbook(blob_stream, data_only=True)

        sheets = []
        total_start_time = time.time()
    
        for sheet_name in workbook.sheetnames:
            logging.info(f"[spreadsheet_chunker][{self.filename}][spreadsheet_process][{sheet_name}] Started processing.")                  
            start_time = time.time()
            sheet_dict = {}            
            sheet_dict['name'] = sheet_name
            sheet = workbook[sheet_name]
            data, headers = self._get_sheet_data(sheet)
            sheet_dict["headers"] = headers
            sheet_dict["data"] = data

            # Only build the full-sheet table/summary when chunking by sheet.
            if not self.chunking_by_row:
                table = tabulate(data, headers=headers, tablefmt="grid")
                table = self._clean_markdown_table(table)
                sheet_dict["table"] = table

                prompt = f"Summarize the table with data in it, by understanding the information clearly.\n table_data:{table}"
                summary = self.aoai_client.get_completion(prompt, max_tokens=2048)
                sheet_dict["summary"] = summary
                logging.debug(
                    f"[spreadsheet_chunker][{self.filename}][spreadsheet_process][{sheet_dict['name']}] Generated summary."
                )
            else:
                sheet_dict["table"] = ""
                sheet_dict["summary"] = ""
                logging.debug(
                    f"[spreadsheet_chunker][{self.filename}][spreadsheet_process][{sheet_dict['name']}] "
                    "Skipped table/summary generation (chunking by row)."
                )
            
            elapsed_time = time.time() - start_time
            logging.debug(f"[spreadsheet_chunker][{self.filename}][spreadsheet_process][{sheet_dict['name']}] Processed in {elapsed_time:.2f} seconds.")
            sheets.append(sheet_dict)
    
        total_elapsed_time = time.time() - total_start_time
        logging.debug(f"[spreadsheet_chunker][{self.filename}][spreadsheet_process] Total processing time: {total_elapsed_time:.2f} seconds.")

        return sheets

    def _row_to_embedding_text(
        self,
        headers,
        row,
        sheet_name,
        row_index,
        include_header_in_embedding,
    ):
        """Build a compact per-row text for embeddings.

        Goals:
        - Dramatically smaller than markdown tables (lower TPM pressure).
        - If include_header_in_embedding=True, ensure the header schema is present in the embedding.

        Format (when include_header_in_embedding=True):
            file=<filename>\n
            sheet=<sheet>\n
            row=<row_index>\n
            cols=<h1>|<h2>|...\n
            vals=<v1>|<v2>|...
        """

        def _norm(value: object) -> str:
            if value is None:
                return ""
            text = str(value)
            text = " ".join(text.replace("\r", " ").replace("\n", " ").split())
            return text.strip()

        safe_headers = [_norm(h) for h in (headers or [])]
        safe_row = [_norm(v) for v in (row or [])]

        width = max(len(safe_headers), len(safe_row))
        if len(safe_headers) < width:
            safe_headers += [""] * (width - len(safe_headers))
        if len(safe_row) < width:
            safe_row += [""] * (width - len(safe_row))

        # Keep alignment; but drop trailing fully-empty columns.
        last_nonempty = -1
        for i, (h, v) in enumerate(zip(safe_headers, safe_row)):
            if h or v:
                last_nonempty = i
        if last_nonempty >= 0:
            safe_headers = safe_headers[: last_nonempty + 1]
            safe_row = safe_row[: last_nonempty + 1]

        # We keep empty values to preserve positional alignment between cols and vals.
        # To reduce tokens, use tight separators and collapsed whitespace.
        cols = "|".join(safe_headers)
        vals = "|".join(safe_row)

        parts = [
            f"file={_norm(self.filename)}",
            f"sheet={_norm(sheet_name)}",
            f"row={row_index}",
        ]
        if include_header_in_embedding:
            parts.append(f"cols={cols}")
        parts.append(f"vals={vals}")
        return "\n".join(parts)

    def _get_sheet_data(self, sheet):
        """
        Retrieves data and headers from the given sheet. Each row's data is processed into a list format, ensuring that empty rows are excluded.

        Args:
            sheet (Worksheet): The worksheet object to extract data from.

        Returns:
            Tuple[List[List[str]], List[str]]: A tuple containing a list of row data and a list of headers.
        """
        data = []
        for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip headers
            row_data = []
            for cell in row:
                cell_value = cell.value
                if cell_value is None:
                    cell_value = ""
                cell_text = str(cell_value)
                row_data.append(cell_text)
            if "".join(row_data).strip() != "":
                data.append(row_data)

        headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]
        return data, headers
    
    def _clean_markdown_table(self, table_str):
        """
        Cleans up a Markdown table string by removing excessive whitespace from each cell.

        Args:
            table_str (str): The Markdown table string to be cleaned.

        Returns:
            str: The cleaned Markdown table string with reduced whitespace.
        """
        cleaned_lines = []
        lines = table_str.splitlines()

        for line in lines:
            if set(line.strip()) <= set('-| '):
                cleaned_lines.append(line)
                continue

            cells = line.split('|')
            stripped_cells = [cell.strip() for cell in cells[1:-1]]
            cleaned_line = '| ' + ' | '.join(stripped_cells) + ' |'
            cleaned_lines.append(cleaned_line)

        return '\n'.join(cleaned_lines)